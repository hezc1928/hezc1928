from __future__ import division

import openpyxl
import sys
import time

import numpy as np
import pandas as pd

# 调用该函数求出lnZ
# 只考虑p1和p4
import lnZ
import matplotlib.pyplot as plt
from Han import BAG, CRU,  TC

import xlrd
import xlwt
sys.path.append('../')
PRE = 10
RANGE = 20
TMIN = 150
TMAX = 200


def initialize():
    # 画布初始化
    ax = plt.gca()
    plt.grid(visible=None, which='major', axis='both',
             color='k', linestyle='--', linewidth=0.5)
    # ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.02f GeV'))
    # 标签设置
    # plt.title('$\Lambda_2 = 3$')
    # plt.tick_params(axis='both', which='major', labelsize=15)
    # plt.xlabel('$\Lambda_1$', fontsize=15)
    # plt.ylabel('$p_i$', fontsize=15)
    # 坐标轴尺度
    # plt.xlim(0.9, 1.1)
    # plt.ylim(0, 1)
    # 坐标轴刻度
    # x_major_locator = mticker.MultipleLocator(1)
    # y_major_locator = mticker.MultipleLocator(1)
    # ax.xaxis.set_major_locator(x_major_locator)
    # ax.yaxis.set_major_locator(y_major_locator)


def cal_QGP(T, L1, L2, num_u=1, num_s=1, num_d=1,  num_g=1, plot=1):
    lnZ_g_up = lnZ.main_list(T=T, MI=0, GI=1, L1=L1, L2=L2)  # gluon
    lnZ_u_up = lnZ.main_list(T=T, MI=0, GI=1/2, L1=L1, L2=L2)  # u quark
    lnZ_d_up = lnZ_u_up  # d quark
    lnZ_s_up = lnZ.main_list(T=T, MI=150, GI=1/2, L1=L1, L2=L2)  # s quark

    lnZ_g_down = lnZ.main(T=TC, MI=0, GI=1, L1=1/CRU, L2=1/CRU)  # gluon
    lnZ_u_down = lnZ.main(T=TC, MI=0, GI=1/2, L1=1/CRU, L2=1/CRU)  # u quark
    lnZ_d_down = lnZ_u_down  # d quark
    lnZ_s_down = lnZ.main(T=TC, MI=150, GI=1/2, L1=1/CRU, L2=1/CRU)  # s quark
    omega_up = []
    omega_down = []
    r = []
    print("1")
    for i in range(len(T)):
        omega_up.append(-T[i]*(num_d*lnZ_d_up[i]+num_g*lnZ_g_up[i] +
                        num_u*lnZ_u_up[i]+num_s*lnZ_s_up[i])+BAG)
        omega_down.append(-T[i]*(num_d*lnZ_d_down +
                          num_g*lnZ_g_down+num_u*lnZ_u_down+num_s*lnZ_s_down)+BAG)
        r.append(omega_up[i]/omega_down[i])
    if num_u == 1:
        plt.plot(T, r, label="u-quark")
    if num_d == 1:
        plt.plot(T, r, label="d-quark")
    if num_s == 1:
        plt.plot(T, r, label="s-quark")
    if num_g == 1:
        plt.plot(T, r, label="gluon")
    return r


def cal_HG_file(T, L1, L2, plot=1):
    a = read_file_HG()
    omega_up = []
    omega_down = []
    lnZ_HG_up = np.zeros(len(T))  # 记得初始化
    lnZ_HG_down = 0
    r = []
    for i in range(len(a)):
        lnZ_up = lnZ.main_list(
            T=T, MI=a[i][0], GI=a[i][3], L1=L1, L2=L2)  # lnZ是list类型
        for j in range(len(lnZ_up)):
            lnZ_up[j] = lnZ_up[j] * a[i][6] * (a[i][5]+1)
        # lnZ_up = [lnZ_up[j] * a[i][6] for j in range(len(lnZ_up))]
        lnZ_HG_up = lnZ_up+lnZ_HG_up

        lnZ_down = lnZ.main(
            T=TC, MI=a[i][0], GI=a[i][3], L1=1/CRU, L2=1/CRU) * a[i][6] * (a[i][5]+1)
        lnZ_HG_down = lnZ_down + lnZ_HG_down

        print(i/len(a))
    for i in range(len(T)):
        omega_up.append(-T[i]*(lnZ_HG_up[i]))  # HG
        omega_down.append(-T[i]*(lnZ_HG_down))  # HG
        r.append(omega_up[i]/omega_down[i])
    if plot == 1:
        plt.plot(T, r, label="HG")
    return r


def cal_HG_file_var(T, L, R, plot=1):
    L1 = 2/(1/L+1/R+CRU)
    L2 = 2/(1/L-1/R+CRU)
    a = read_file_HG()
    omega_up = []
    omega_down = []
    lnZ_HG_up = np.zeros(len(T))  # 记得初始化
    lnZ_HG_down = 0
    r = []
    for i in range(len(a)):
        lnZ_up = lnZ.main_list(
            T=T, MI=a[i][0], GI=a[i][3], L1=L1, L2=L2)  # lnZ是list类型
        for j in range(len(lnZ_up)):
            lnZ_up[j] = lnZ_up[j] * a[i][6] * (a[i][5]+1)
        # lnZ_up = [lnZ_up[j] * a[i][6] for j in range(len(lnZ_up))]
        lnZ_HG_up = lnZ_up+lnZ_HG_up

        lnZ_down = lnZ.main(
            T=TC, MI=a[i][0], GI=a[i][3], L1=1/CRU, L2=1/CRU) * a[i][6] * (a[i][5]+1)
        lnZ_HG_down = lnZ_down + lnZ_HG_down

        print(i/len(a))
    for i in range(len(T)):
        omega_up.append(-T[i]*(lnZ_HG_up[i]))  # HG
        omega_down.append(-T[i]*(lnZ_HG_down))  # HG
        r.append(omega_up[i]/omega_down[i])
    if plot == 1:
        plt.plot(T, r, label="HG")
    return r


def cal_QGP_file(T, L1, L2, plot=1):
    a = read_file_QGP()
    omega_up = []
    omega_down = []
    lnZ_QGP_up = np.zeros(len(T))  # 记得初始化
    lnZ_QGP_down = 0
    r = []
    for i in range(len(a)):
        lnZ_up = lnZ.main_list(
            T=T, MI=a[i][0], GI=a[i][3], L1=L1, L2=L2)  # lnZ是list类型
        for j in range(len(lnZ_up)):
            lnZ_up[j] = lnZ_up[j] * a[i][6] * (a[i][5]+1)
        # lnZ_up = [lnZ_up[j] * a[i][6] for j in range(len(lnZ_up))]
        lnZ_QGP_up = lnZ_up+lnZ_QGP_up

        lnZ_down = lnZ.main(
            T=TC, MI=a[i][0], GI=a[i][3], L1=1/CRU, L2=1/CRU) * a[i][6] * (a[i][5]+1)
        lnZ_QGP_down = lnZ_down + lnZ_QGP_down

        print(i/len(a))
    for i in range(len(T)):
        omega_up.append(-T[i]*(lnZ_QGP_up[i])+BAG)
        omega_down.append(-T[i]*(lnZ_QGP_down)+BAG)
        r.append(omega_up[i]/omega_down[i])
    if plot == 1:
        plt.plot(T, r, label="QGP")
    return r


def cal_QGP_file_var(T, L, R, plot=1):
    L1 = 2/(1/L+1/R+CRU)
    L2 = 2/(1/L-1/R+CRU)
    a = read_file_QGP()
    omega_up = []
    omega_down = []
    lnZ_QGP_up = np.zeros(len(T))  # 记得初始化
    lnZ_QGP_down = 0
    r = []
    for i in range(len(a)):
        lnZ_up = lnZ.main_list(
            T=T, MI=a[i][0], GI=a[i][3], L1=L1, L2=L2)  # lnZ是list类型
        for j in range(len(lnZ_up)):
            lnZ_up[j] = lnZ_up[j] * a[i][6] * (a[i][5]+1)
        # lnZ_up = [lnZ_up[j] * a[i][6] for j in range(len(lnZ_up))]
        lnZ_QGP_up = lnZ_up+lnZ_QGP_up

        lnZ_down = lnZ.main(
            T=TC, MI=a[i][0], GI=a[i][3], L1=1/CRU, L2=1/CRU) * a[i][6] * (a[i][5]+1)
        lnZ_QGP_down = lnZ_down + lnZ_QGP_down

        print(i/len(a))
    for i in range(len(T)):
        omega_up.append(-T[i]*(lnZ_QGP_up[i])+BAG)
        omega_down.append(-T[i]*(lnZ_QGP_down)+BAG)
        r.append(omega_up[i]/omega_down[i])
    if plot == 1:
        plt.plot(T, r, label="QGP")
    return r


def plot_lnZ_HG(T, L1, L2, plot=1):
    a = read_file_HG()
    omega_up = []
    lnZ_HG_up = np.zeros(len(T))  # 记得初始化
    omega_down = []
    lnZ_QGP_down = 0
    r = []
    for i in range(len(a)):
        # lnZ_up = lnZ.main_list(
        #     T=T, MI=a[i][0], GI=a[i][3], L1=L1, L2=L2)  # lnZ是list类型
        # for j in range(len(lnZ_up)):
        #     lnZ_up[j] = lnZ_up[j] * a[i][6] *( a[i][5]+1)
        # # lnZ_up = [lnZ_up[j] * a[i][6] for j in range(len(lnZ_up))]
        # lnZ_HG_up = lnZ_up+lnZ_HG_up

        lnZ_down = lnZ.main(
            T=TC, MI=a[i][0], GI=a[i][3], L1=1/CRU, L2=1/CRU) * a[i][6] * (a[i][5]+1)
        lnZ_QGP_down = lnZ_down + lnZ_QGP_down

        print(i/len(a))
    for i in range(len(T)):
        # omega_up.append(-T[i]*(lnZ_HG_up[i]))
        # omega_up.append((lnZ_HG_up[i]))
        omega_down.append(-T[i]*(lnZ_QGP_down))
        # r.append(omega_up[i]/omega_down[i])
    if plot == 1:
        plt.plot(T, omega_down, label="HG")
    return r


def plot_lnZ_QGP(T, L1, L2, plot=1):
    a = read_file_QGP()
    omega_up = []
    lnZ_HG_up = np.zeros(len(T))  # 记得初始化
    omega_down = []
    lnZ_QGP_down = 0
    r = []
    for i in range(len(a)):
        # lnZ_up = lnZ.main_list(
        #     T=T, MI=a[i][0], GI=a[i][3], L1=L1, L2=L2)  # lnZ是list类型
        # for j in range(len(lnZ_up)):
        #     lnZ_up[j] = lnZ_up[j] * a[i][6] *( a[i][5]+1)
        # # lnZ_up = [lnZ_up[j] * a[i][6] for j in range(len(lnZ_up))]
        # lnZ_HG_up = lnZ_up+lnZ_HG_up

        lnZ_down = lnZ.main(
            T=TC, MI=a[i][0], GI=a[i][3], L1=1/CRU, L2=1/CRU) * a[i][6] * (a[i][5]+1)
        lnZ_QGP_down = lnZ_down + lnZ_QGP_down

        print(i/len(a))
    for i in range(len(T)):
        # omega_up.append(-T[i]*(lnZ_HG_up[i])+BAG)
        # omega_up.append((lnZ_HG_up[i]))
        omega_down.append(-T[i]*(lnZ_QGP_down)+BAG)
        # r.append(omega_up[i]/omega_down[i])
    if plot == 1:
        plt.plot(T, omega_down, label="QGP")
    return r


def read_file_HG():
    df = pd.read_csv('E:\cpp_project\py\matplotlib\part_table_HG_test.dat',
                     sep="\t", header=None)
    a = df.values.tolist()
    return a


def read_file_QGP():
    df = pd.read_csv('E:\cpp_project\py\matplotlib\part_table_QGP_test.dat',
                     sep="\t", header=None)
    a = df.values.tolist()
    return a


def find_Tc(r_HG, r_QGP):
    delta_r = []
    for i in range(len(r_HG)):
        delta_r.append(np.abs(r_HG[i]-r_QGP[i]))
    position = delta_r.index(min(delta_r))
    Tc = TMIN+position*1/PRE
    return Tc


def Tc_L_plot(T, L):
    Tc = []

    data_QGP = openpyxl.Workbook()
    data_QGP.create_sheet('Sheet1')
    table_QGP = data_QGP.active

    data_HG = openpyxl.Workbook()
    data_HG.create_sheet('Sheet1')
    table_HG = data_HG.active

    data_Tc = openpyxl.Workbook()
    data_Tc.create_sheet('Sheet1')
    table_Tc = data_Tc.active
    for i in range(len(L)):
        r_HG = cal_HG_file_var(T, L[i]/2, 1/CRU, 0)
        xls_write(r_HG, i, table_QGP)
        r_QGP = cal_QGP_file_var(T, L[i]/2, 1/CRU, 0)
        xls_write(r_QGP, i, table_HG)
        Tc.append(find_Tc(r_HG=r_HG, r_QGP=r_QGP))
    xls_write(Tc, 1, table_Tc)
    data_QGP.save("E:\\cpp_project\\py\\matplotlib\\data\\QGP.xlsx")
    data_HG.save("E:\\cpp_project\\py\\matplotlib\\data\\HG.xlsx")
    data_Tc.save("E:\\cpp_project\\py\\matplotlib\\data\\Tc.xlsx")
    plt.plot(L, Tc, label="final")


def xls_write(var, col, table):
    for count in range(len(var)):
        table.cell(count+1, col+1, var[count])


# def xls_write_QGP(var, col):
#     for count in range(len(var)):
#         table_QGP.cell(count+1, col+1, var[count])


# def xls_write_Tc(var, col):
#     for count in range(len(var)):
#         table_Tc.cell(count+1, col+1, var[count])


# def xls_write_HG(var, col):
#     for count in range(len(var)):
#         table_HG.cell(count+1, col, var[count])


def xls_write_test():
    data = openpyxl.Workbook()  # 新建工作簿
    data.create_sheet('Sheet1')  # 添加页
    # table = data.get_sheet_by_name('Sheet1') # 获得指定名称页
    table = data.active  # 获得当前活跃的工作页，默认为第一个工作页
    table.cell(1, 2, 'Test')  # 行，列，值 这里是从1开始计数的
    data.save("E:\\cpp_project\\py\\matplotlib\\data\\test.xlsx")  # 一定要保存


def main():

    start = time.time()
    initialize()
    T = np.arange(TMIN, TMAX, 1/PRE)
    # L = np.arange(2, 9, 1/PRE)
    # L = [1/CRU]
    # Tc_L_plot(T, L)

    # cal_QGP(T, 4, 4, 0, 0, 0, 16)
    # cal_QGP(T, 4, 4, 0, 6, 0, 0)
    # cal_QGP(T, 4, 4, 6, 0, 0, 0)
    # cal_QGP(T, 4, 4, 0, 0, 0, 1)
    # cal_QGP(T, 4, 4, 0, 1, 0, 0)
    # cal_QGP(T, 4, 4, 1, 0, 0, 0)
    plot_lnZ_HG(T, 4, 4, 1)
    plot_lnZ_QGP(T, 4, 4, 1)
    plot_lnZ_HG(T, 1/CRU, 1/CRU, 1)
    plot_lnZ_QGP(T, 1/CRU, 1/CRU, 1)
    # cal_QGP_file(T, 4, 1/CRU)
    # cal_HG_file(T, 4, 1/CRU)
    # cal_QGP_file(T, 1/CRU, 1/CRU)
    # cal_HG_file(T, 1/CRU, 1/CRU)
    end = time.time()
    print("time = ", end-start)
    plt.legend()
    plt.show()
    return 0


if __name__ == '__main__':
    main()
